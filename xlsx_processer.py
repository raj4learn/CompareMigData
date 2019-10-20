# xlsx processer
# This program will be used to process the xlsx file.
from openpyxl import load_workbook
import traceback as err
import itertools
import pyexcelerate as pyxl
from openpyxl.styles import Font, PatternFill, colors, Border, Side, Alignment, Protection
from openpyxl.utils import get_column_letter


# fnReadQuires
# Purpose: This function will have the details of the Mapping sheet content.read_Quires
# With this we will be creating the Source and Destination instance Query
# l_key will help to mark the Key columns, for the comparision

def fnReadQuires(p_fileName, p_sheetName, p_Header = 1):
    one_table_set_d = dict()
    print(f"read_Quires: Read excel file: {p_fileName}:{p_sheetName}")
    try:
        wb = load_workbook(p_fileName, read_only=True, data_only=False)
        ws = wb[p_sheetName]
        l_max_row = ws.max_row
        l_max_col = ws.max_column
        #print(f"Row:{l_max_row} and Col:{l_max_col}")

        if l_max_row <= 0 or l_max_col <=0:
            return dict()

        l_previous_rec = ""
        for r in range(p_Header + 1, l_max_row + 1):
            col_header_l = list()
            for c in range(1, l_max_col + 1):
                cell_obj = ws.cell(row=r, column=c)
                col_header_l.append(cell_obj.value)

            if l_previous_rec != col_header_l[0]:
                if one_table_set_d.__len__() > 0:
                    #print(f"Return Final Table Records In : {one_table_set_d}" )
                    yield one_table_set_d
                    one_table_set_d.clear()
                    #print(f"Alter yield current Data : {one_table_set_d}-{l_previous_rec}~{col_header_l[0]}")
                l_previous_rec = col_header_l[0]

            one_table_set_d[r] = col_header_l
            #print(f"Stored {r}:{col_header_l}")

    except Exception as e:
        print(f"Error in Reading Input Compare Quires List Sheet")
        print(f"{e}")
    else:
        #print(f"Return Final Table Records Out : {one_table_set_d}")
        yield one_table_set_d
    finally:
        print(f"Finally Closing Query Setup WorkBook {p_fileName}")
        wb.close()


def apply_cell_color(p_sheet, p_row, p_col, p_code):
    my_blue = colors.Color(rgb='00003096')
    col_white = colors.Color(rgb='00FFFFFF')
    try:
        if p_code == "HEADER":
            p_sheet.cell(row=p_row, column=p_col).fill = PatternFill(fgColor=my_blue, patternType='solid')
            p_sheet.cell(row=p_row, column=p_col).font = Font(b=True, color=col_white, size=10)
            # print(f" size : {len(p_sheet.cell(row=p_row, column=p_col).value)}")
            try:
                l_size = ( len(p_sheet.cell(row=p_row, column=p_col).value) + 5 )
            except:
                l_size = 40
                # print(f"Warning Unable to get the size. Default to 40")
            p_sheet.column_dimensions[get_column_letter(p_col)].width = l_size
    except:
        print(f"Error in Applying Color[{p_row}:{p_col}]\n Error:{err.format_exc()}")


def write2Xlsx(p_s_d_flag, p_tbl, p_header, p_Key_cols, p_result, p_out_file):
    l_write_flag = 0
    l_ShName = p_s_d_flag + "_" + p_tbl
    print(f"Xlsx Sheet Name: {l_ShName}")
    l_inputFile = p_out_file
    try:
        wb_obj = load_workbook(l_inputFile)
        l_sheet_found = 1 if l_ShName in wb_obj.sheetnames else 0
        try:
            xlsh = wb_obj.get_sheet_by_name(l_ShName) if l_sheet_found == 1 else wb_obj.create_sheet(title=l_ShName)

            if l_sheet_found == 0:
                xlsh.sheet_properties.tabColor = "1072BA"
                #print(f"Header Length: {p_header.__len__()}")
                if p_header.__len__() > 0:
                    for idx_col, val in enumerate(p_header):
                        xlsh.cell(1, (idx_col + 1), val)

            l_max_row = xlsh.max_row
            if l_max_row == 1:
                #print(f" first row value: {xlsh.cell(l_max_row, 1).value}")
                if xlsh.cell(l_max_row, 1).value is None:
                    l_max_row = 0

            #print(f"Currnet Row OUT: {xlsh.max_row}-{l_max_row}")

            for row in p_result:
                #print(f"write2Xlsx: row: {row}")
                l_col = 1
                l_max_row = l_max_row + 1
                #print(f"Currnet Row: {l_max_row}")
                for col in row:
                    xlsh.cell(l_max_row, l_col, col)

                    # Apply the Stype after the Value is populated.
                    if l_max_row == 1:
                        apply_cell_color(xlsh, l_max_row, l_col, 'HEADER')
                    l_col += 1

            l_write_flag = 1
        except Exception as e:
            print(f"writeXlsx: Error at Level-2: {e}")
        finally:
            if (l_write_flag == 1):
                print(f"Saving the File.")
                wb_obj.save(l_inputFile)
    except Exception as e:
        print(f"writeXlsx: Error at Level-0: {err.format_exc()}")
        return False
    else:
        return True
    finally:
        if wb_obj is not None:
            wb_obj.close()
            print(f"Finally Closing the Workbook [{p_out_file}]")


# fnBuildQuery
# This function help to Create a query from the Excel sheet shared data.
# Function returns Source and Destination
def fnBuildQuery(p_tblDataFromXlsxD):
    l_s_Key_col = l_s_col_nm = l_s_tbl_nm = ""
    l_d_tbl_nm = l_d_col_nm = l_d_Key_col = ""

    for lk, lv in p_tblDataFromXlsxD.items():
        # Source Table Name
        l_s_tbl_nm = lv[0] if (lv[0] is not None and l_s_tbl_nm == "") else l_s_tbl_nm
        # Destination Table Name
        l_d_tbl_nm = lv[3] if (lv[3] is not None and l_d_tbl_nm == "") else l_d_tbl_nm

        # Source Column Names
        if lv[1] is not None:
            l_s_col_nm = l_s_col_nm + lv[1] if (l_s_col_nm == "") else (l_s_col_nm + "," + lv[1])
        # Destination Column Names
        if lv[4] is not None:
            l_d_col_nm = l_d_col_nm + lv[4] if (l_d_col_nm == "") else (l_d_col_nm + "," + lv[4])

        # Source Key
        if str(lv[2]).upper() == 'YES':
            l_s_Key_col = l_s_Key_col + lv[1] if (l_s_Key_col == "") else (l_s_Key_col + "," + lv[1])

        # Destination Key
        if str(lv[5]).upper() == 'YES':
            l_d_Key_col = l_d_Key_col + lv[4] if (l_d_Key_col == "") else (l_d_Key_col + "," + lv[4])

    #print(f"fnBuildQuery Source Query: {l_s_tbl_nm}:{l_s_col_nm}:{l_s_Key_col}")
    #print(f"fnBuildQuery Destination Query: {l_d_tbl_nm}:{l_d_col_nm}:{l_d_Key_col}")
    l_s_sql = "SELECT " + l_s_col_nm + " FROM " + l_s_tbl_nm
    l_d_sql = "SELECT " + l_d_col_nm + " FROM " + l_d_tbl_nm

    return l_s_tbl_nm, l_d_tbl_nm, l_s_sql, l_d_sql, l_s_Key_col, l_d_Key_col


# This main function read Master Sheet in Excel and selects the tables for comparision..
def read_proc_filter(p_excel_input_file, p_sheet_name, p_header_row, p_module, p_form, p_table):
    l_tables = dict()
    wb = load_workbook(p_excel_input_file, read_only=True, data_only=False)
    wb_sh = wb[p_sheet_name]
    l_max_row = wb_sh.max_row
    l_max_col = wb_sh.max_column
    # print(f"Row:{l_max_row} and Col:{l_max_col}")
    rows = wb_sh.rows
    process_row = list()
    for idx, row in enumerate(rows):
        row_tmp = [ row[0].value, row[1].value, row[2].value, row[3].value, ]
        if ( (p_module == row_tmp[0] or  p_module == "ALL") \
                and (p_form == row_tmp[1] or p_form == "ALL") \
                and ( p_table == row_tmp[2] or p_table == "ALL") \
                and ( row_tmp[3] == 'Y' )
            ):
            process_row.append(row_tmp)
            #print(f"idx:{idx} row: {row[0].value}")

    for rec in process_row:
        l_tables[rec[2]] = [rec[0], rec[1], ]

    #print(f"rec: {l_tables}")

    return l_tables



# This main function which call two method in this file.
def read_quires(p_excel_input_file, p_sheet_name, p_header_row):
    for l_key in fnReadQuires(p_excel_input_file, p_sheet_name, p_header_row):
        l_s_tbl, l_d_tbl, lSrcQuery, lDestQuery, lSrcKeys, lDestKeys = fnBuildQuery(l_key)
        #print(f"Main: Source Query: {lSrcQuery}:{lSrcKeys}")
        #print(f"Main: Destination Query: {lDestQuery}:{lDestKeys}")
        #print("******************************************")
        yield l_s_tbl, l_d_tbl, lSrcQuery, lDestQuery, lSrcKeys, lDestKeys

################################################
# Tester
if __name__ == "__main__":
    read_quires("CompareQuiresList.xlsx", "TableList", 1)

################## End of File #################

